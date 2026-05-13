from __future__ import annotations

from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO
from uuid import uuid4

from openpyxl import load_workbook


def test_payments_to_xlsx_produces_valid_workbook(db_session):
    from app.models import Analyst, AnalystStatus, Payment, PaymentStatus, Project, ProjectStatus
    from app.services.exports import payments_to_xlsx

    a = Analyst(
        id=uuid4(),
        full_name="Иван Петров",
        email="i@example.com",
        amo_user_id=10,
        default_rate=Decimal("0"),
        payment_details={"recipient": "Иван Петров", "bank": "Sber", "bik": "044525225", "account": "40817810099910004312"},
        status=AnalystStatus.active,
    )
    db_session.add(a)
    db_session.flush()

    p = Project(
        id=uuid4(),
        amo_deal_id=42,
        name="Внедрение CRM",
        analyst_id=a.id,
        payment_amount=Decimal("150000"),
        status=ProjectStatus.done,
    )
    db_session.add(p)
    db_session.flush()

    pay = Payment(
        id=uuid4(),
        project_id=p.id,
        analyst_id=a.id,
        amount=Decimal("150000"),
        status=PaymentStatus.ready,
        accrued_at=datetime(2026, 5, 1, 12, 0, tzinfo=timezone.utc),
        comment="за май",
    )
    db_session.add(pay)
    db_session.commit()

    blob = payments_to_xlsx(db_session, [pay])
    assert blob.startswith(b"PK")  # zip-сигнатура xlsx

    wb = load_workbook(BytesIO(blob))
    ws = wb.active
    assert ws.title == "Выплаты"
    # шапка
    assert [c.value for c in ws[1]] == [
        "Дата начисления", "Аналитик", "Проект", "Сумма", "Статус", "Реквизиты", "Комментарий",
    ]
    # данные
    row = [c.value for c in ws[2]]
    assert row[1] == "Иван Петров"
    assert row[2] == "Внедрение CRM"
    assert row[3] == 150000.0
    assert row[4] == "К выплате"
    assert "Sber" in row[5]
    assert "044525225" in row[5]
    assert row[6] == "за май"
