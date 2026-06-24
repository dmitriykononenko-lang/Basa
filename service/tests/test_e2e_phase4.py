"""End-to-end Фазы 4: XLSX-экспорт, алерты, SPA-mount."""

from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from uuid import UUID, uuid4

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.dialects import postgresql
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool
from sqlalchemy.types import JSON


@pytest.fixture()
def env(monkeypatch):
    from app.db.base import Base
    import app.models  # noqa: F401
    from app.core.security import hash_password
    from app.models import User, UserRole

    for table in Base.metadata.tables.values():
        for col in table.columns:
            if isinstance(col.type, postgresql.JSONB):
                col.type = JSON()
            if col.server_default is not None and "gen_random_uuid" in str(col.server_default.arg):
                col.server_default = None
                if col.default is None and col.primary_key:
                    col.default = lambda: uuid4()

    engine = create_engine(
        "sqlite:///:memory:",
        future=True,
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)

    from app.db import session as session_mod
    session_mod.engine = engine
    session_mod.SessionLocal = Session

    db = Session()
    admin = User(email="admin@example.com", password_hash=hash_password("admin12345"), role=UserRole.admin, is_active=True)
    accountant = User(email="acc@example.com", password_hash=hash_password("acc12345"), role=UserRole.accountant, is_active=True)
    analyst_user = User(email="ann@example.com", password_hash=hash_password("ann12345"), role=UserRole.analyst, is_active=True)
    db.add_all([admin, accountant, analyst_user])
    db.commit()

    from app.main import app

    def _get_db():
        try:
            yield db
        finally:
            pass

    app.dependency_overrides[session_mod.get_db] = _get_db

    # Алерты — на пустую заглушку, чтобы тесты не лезли в Redis
    from app.services import alerts as alerts_mod
    monkeypatch.setattr(alerts_mod, "_redis", lambda: None)

    yield {"client": TestClient(app), "db": db, "users": {"admin": admin, "accountant": accountant, "analyst": analyst_user}}
    db.close()
    app.dependency_overrides.clear()


def _login(client, email, password):
    r = client.post("/api/v1/auth/login", json={"email": email, "password": password})
    assert r.status_code == 200, r.text
    return r.json()["access_token"]


def _hdr(t):
    return {"Authorization": f"Bearer {t}"}


def test_xlsx_export_returns_workbook(env):
    client = env["client"]
    db = env["db"]
    admin_token = _login(client, "admin@example.com", "admin12345")

    # Создаём аналитика и выплату
    from decimal import Decimal
    from app.models import Analyst, AnalystStatus, Payment, PaymentStatus, Project, ProjectStatus

    a = Analyst(
        id=uuid4(), full_name="A", email="a@example.com", amo_user_id=1,
        default_rate=Decimal("1000"), payment_details={"bank": "Sber"}, status=AnalystStatus.active,
    )
    db.add(a)
    db.flush()
    p = Project(id=uuid4(), name="P", analyst_id=a.id, payment_amount=Decimal("1000"), status=ProjectStatus.done)
    db.add(p)
    db.flush()
    pay = Payment(
        id=uuid4(), project_id=p.id, analyst_id=a.id, amount=Decimal("1000"),
        status=PaymentStatus.ready, accrued_at=datetime(2026, 5, 1, tzinfo=timezone.utc),
    )
    db.add(pay)
    db.commit()

    r = client.get("/api/v1/payments/export.xlsx?status=ready", headers=_hdr(admin_token))
    assert r.status_code == 200, r.text
    assert "spreadsheetml" in r.headers["content-type"]
    assert "attachment" in r.headers.get("content-disposition", "")

    wb = load_workbook(BytesIO(r.content))
    ws = wb.active
    assert ws["B2"].value == "A"
    assert ws["D2"].value == 1000.0
    assert ws["E2"].value == "К выплате"


def test_xlsx_export_forbidden_for_analyst(env):
    client = env["client"]
    tok = _login(client, "ann@example.com", "ann12345")
    r = client.get("/api/v1/payments/export.xlsx", headers=_hdr(tok))
    assert r.status_code == 403


def test_alerts_status_admin_only(env):
    client = env["client"]
    admin_token = _login(client, "admin@example.com", "admin12345")
    r = client.get("/api/v1/alerts/status", headers=_hdr(admin_token))
    assert r.status_code == 200
    body = r.json()
    assert body["errors_last_hour"] == 0
    assert body["triggered"] is False

    # accountant — не должен иметь доступ
    acc = _login(client, "acc@example.com", "acc12345")
    r = client.get("/api/v1/alerts/status", headers=_hdr(acc))
    assert r.status_code == 403


def test_spa_root_serves_index_html(env):
    r = env["client"].get("/")
    assert r.status_code == 200
    assert "Basa" in r.text
    assert "app.js" in r.text  # подключается SPA-скрипт


def test_spa_static_assets_served(env):
    r = env["client"].get("/styles.css")
    assert r.status_code == 200
    assert "login-screen" in r.text or "topbar" in r.text  # просто что-то из CSS

    r = env["client"].get("/app.js")
    assert r.status_code == 200
    assert "renderProjects" in r.text


def test_api_routes_not_shadowed_by_spa(env):
    """SPA-mount на / не должен перекрывать /api/v1."""
    r = env["client"].get("/healthz")
    assert r.status_code == 200
    assert r.json()["status"] == "ok"

    # API ещё на месте и не отдаёт SPA
    admin_token = _login(env["client"], "admin@example.com", "admin12345")
    r = env["client"].get("/api/v1/auth/me", headers=_hdr(admin_token))
    assert r.status_code == 200
    assert r.json()["email"] == "admin@example.com"
