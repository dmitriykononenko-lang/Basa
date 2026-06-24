"""Экспорты для бухгалтерии."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models import Analyst, Payment, Project


def payments_to_xlsx(db: Session, payments: Iterable[Payment]) -> bytes:
    """Реестр выплат в XLSX — формат, который без правок принимает 1С/банковский экспорт.

    Колонки: Дата начисления | Аналитик | Проект | Сумма | Статус | Реквизиты | Комментарий.
    Реквизиты подтягиваются из `analysts.payment_details` (jsonb) — строкой через запятую.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Выплаты"

    headers = ["Дата начисления", "Аналитик", "Проект", "Сумма", "Статус", "Реквизиты", "Комментарий"]
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")

    # ширины
    widths = [20, 30, 40, 14, 14, 40, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 2
    for p in payments:
        analyst = db.get(Analyst, p.analyst_id)
        project = db.get(Project, p.project_id)
        ws.cell(row=row, column=1, value=_fmt_dt(p.accrued_at))
        ws.cell(row=row, column=2, value=analyst.full_name if analyst else "")
        ws.cell(row=row, column=3, value=project.name if project else "")
        ws.cell(row=row, column=4, value=float(p.amount))
        ws.cell(row=row, column=4).number_format = "#,##0.00"
        ws.cell(row=row, column=5, value=_status_ru(p.status.value if hasattr(p.status, "value") else str(p.status)))
        ws.cell(row=row, column=6, value=_format_details(analyst.payment_details if analyst else None))
        ws.cell(row=row, column=7, value=p.comment or "")
        row += 1

    # автофильтр на все колонки
    if row > 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row - 1}"
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _fmt_dt(dt: datetime | None) -> str:
    if dt is None:
        return ""
    return dt.strftime("%Y-%m-%d %H:%M")


_STATUS_RU = {
    "accrued": "Начислено",
    "ready": "К выплате",
    "paid": "Выплачено",
    "cancelled": "Отменено",
}


def _status_ru(status: str) -> str:
    return _STATUS_RU.get(status, status)


def _format_details(details: dict | None) -> str:
    if not details or not isinstance(details, dict):
        return ""
    # порядок: ФИО, банк, БИК, счёт, ИНН, другое
    order = ["recipient", "bank", "bik", "account", "inn"]
    parts: list[str] = []
    for key in order:
        if key in details and details[key]:
            parts.append(f"{key}: {details[key]}")
    # хвост — остальные ключи
    for key, value in details.items():
        if key in order or value in (None, ""):
            continue
        parts.append(f"{key}: {value}")
    return "; ".join(parts)
